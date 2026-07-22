"""Xiaocan (小蚕) Excel parsing + MySQL import + matching service.

Imports xiaocan orders into xiaocan_orders table (truncate-and-replace semantics),
then updates coffee_orders.is_xiaocan / updated_at by matching
xiaocan_orders.platform_order_no = coffee_orders.order_no.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import Any, Iterator

from openpyxl import load_workbook

from ..models import XiaocanImportResult

logger = logging.getLogger(__name__)

# Field -> acceptable header names (matches the actual xiaocan Excel layout).
FIELD_ALIASES: dict[str, list[str]] = {
    "xiaocan_order_no": ["小蚕订单编号", "订单编号", "小蚕订单号", "xiaocan_order_no"],
    "platform": ["下单平台", "平台", "platform"],
    "order_time": ["订单时间", "下单时间", "order_time"],
    "platform_order_no": ["平台订单编号", "平台订单号", "外卖平台单号", "支付平台单号", "平台单号", "platform_order_no"],
    "settlement_amount": ["结算金额", "settlement_amount"],
}

NUMERIC_FIELDS = {"settlement_amount"}
DATE_FIELDS = {"order_time"}

INSERT_SQL = """
INSERT INTO xiaocan_orders (
  xiaocan_order_no, platform, order_time, platform_order_no, settlement_amount
) VALUES (
  %(xiaocan_order_no)s, %(platform)s, %(order_time)s, %(platform_order_no)s, %(settlement_amount)s
)
"""

_PLACEHOLDERS = {"", "-", "--", "none", "null", "n/a", "na"}


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip().lower() in _PLACEHOLDERS
    return False


def _normalize_text(v: Any) -> str | None:
    if _is_blank(v):
        return None
    return str(v).strip()


def _normalize_numeric(v: Any) -> Decimal | None:
    if _is_blank(v):
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    try:
        return Decimal(str(v).strip())
    except (ValueError, ArithmeticError):
        return None


def _normalize_date(v: Any) -> datetime | None:
    if _is_blank(v):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        try:
            return datetime.fromordinal(int(v) + 693594)
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _build_header_map(header_row: Any) -> dict[str, int]:
    headers: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        raw = cell.value
        if raw is None:
            continue
        headers[idx] = str(raw).strip()

    field_to_col: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        for col, hname in headers.items():
            if hname in aliases:
                field_to_col[field] = col
                break

    logger.info("Xiaocan Excel headers detected: %s", headers)
    logger.info("Xiaocan field-to-column mapping: %s", field_to_col)
    return field_to_col


def _extract_values(row: Any, field_to_col: dict[str, int]) -> dict[str, Any]:
    cells = [cell.value for cell in row]
    values: dict[str, Any] = {}
    for field, col in field_to_col.items():
        values[field] = cells[col] if col < len(cells) else None
    return values


def _build_record(values: dict[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for field in FIELD_ALIASES:
        raw = values.get(field)
        if field in NUMERIC_FIELDS:
            record[field] = _normalize_numeric(raw)
        elif field in DATE_FIELDS:
            record[field] = _normalize_date(raw)
        else:
            record[field] = _normalize_text(raw)
    return record


def _iter_records(file_bytes: bytes) -> Iterator[dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=False)

        try:
            header_row = next(row_iter)
        except StopIteration:
            return

        field_to_col = _build_header_map(header_row)
        if not field_to_col:
            raise ValueError(
                "未能识别任何表头列，请检查 Excel 第一行是否为：订单编号 / 下单平台 / 订单时间 / 平台订单编号 / 结算金额"
            )

        for row in row_iter:
            values = _extract_values(row, field_to_col)
            if all(_is_blank(v) for v in values.values()):
                continue
            yield _build_record(values)
    finally:
        wb.close()


def update_xiaocan_marks(conn: Any) -> tuple[int, int]:
    """Recompute coffee_orders.is_xiaocan by matching against xiaocan_orders.

    Efficient approach:
    Step 1: Reset is_xiaocan=0 only for rows that were matched but no longer have a match
    Step 2: Set is_xiaocan=1 only for rows that now have a match (using index on platform_order_no)
    Step 3: Count unmatched xiaocan_orders rows

    Returns (matched_count, unmatched_count).
    """
    cur = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE coffee_orders co
            SET co.is_xiaocan = 0, co.updated_at = NOW()
            WHERE co.is_xiaocan = 1
              AND NOT EXISTS (
                SELECT 1 FROM xiaocan_orders xo
                WHERE co.platform_order_no = COALESCE(NULLIF(TRIM(xo.platform_order_no), ''), xo.xiaocan_order_no)
              )
            """
        )

        cur.execute(
            """
            UPDATE coffee_orders co
            JOIN xiaocan_orders xo
              ON co.platform_order_no = COALESCE(NULLIF(TRIM(xo.platform_order_no), ''), xo.xiaocan_order_no)
            SET co.is_xiaocan = 1,
                co.updated_at = NOW()
            WHERE co.is_xiaocan = 0
              AND COALESCE(NULLIF(TRIM(xo.platform_order_no), ''), xo.xiaocan_order_no) IS NOT NULL
            """
        )
        matched_count = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0

        cur.execute(
            """
            SELECT COUNT(*) FROM xiaocan_orders xo
            WHERE COALESCE(NULLIF(TRIM(xo.platform_order_no), ''), xo.xiaocan_order_no) IS NOT NULL
              AND NOT EXISTS (
                SELECT 1 FROM coffee_orders co
                WHERE co.platform_order_no = COALESCE(NULLIF(TRIM(xo.platform_order_no), ''), xo.xiaocan_order_no)
              )
            """
        )
        unmatched_count = cur.fetchone()[0] or 0

        cur.close()
        return matched_count, unmatched_count
    except Exception:
        cur.close()
        raise


def import_xiaocan_excel(conn: Any, file_bytes: bytes) -> XiaocanImportResult:
    """Truncate xiaocan_orders, import all rows, then refresh is_xiaocan marks."""
    cursor = conn.cursor()
    total_rows = 0
    batch: list[dict[str, Any]] = []
    BATCH_SIZE = 500

    try:
        cursor.execute("TRUNCATE TABLE xiaocan_orders")

        for record in _iter_records(file_bytes):
            batch.append(record)
            total_rows += 1
            if len(batch) >= BATCH_SIZE:
                cursor.executemany(INSERT_SQL, batch)
                batch.clear()

        if batch:
            cursor.executemany(INSERT_SQL, batch)
            batch.clear()

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()

    matched_count, unmatched_count = update_xiaocan_marks(conn)
    conn.commit()

    return XiaocanImportResult(
        total_rows=total_rows,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        truncated=True,
    )
