"""Excel parsing + MySQL import service.

Dynamically maps columns by header name (order-independent), forward-fills
order-level fields for detail rows of the same order, and batch-inserts
into MySQL with full-replace (truncate) semantics.
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Iterator

from openpyxl import load_workbook

from ..models import ImportResult

# Field -> acceptable header names. First match in header row wins.
# Order-independent: whatever column the header sits in, we pick it up.
FIELD_ALIASES: dict[str, list[str]] = {
    "order_no": ["流水号", "订单号", "订单编号", "order_no"],
    "order_date": ["日期", "下单日期", "订单日期", "order_date"],
    "store_name": ["门店", "店铺", "store_name"],
    "queue_no": ["牌号", "取餐号", "queue_no"],
    "order_source": ["渠道", "订单来源", "来源", "order_source"],
    "delivery_method": ["配送方式", "delivery_method"],
    "status": ["状态", "订单状态", "status"],
    "payment_method": ["支付方式", "付款方式", "payment_method"],
    "member_no": ["会员", "会员号", "member_no"],
    "customer_name": ["姓名", "客户姓名", "顾客姓名", "customer_name"],
    "phone": ["电话", "手机号", "phone"],
    "address": ["地址", "收货地址", "address"],
    "product_name": ["商品信息", "商品名称", "商品", "product_name"],
    "flavor_group": ["口味组", "口味", "flavor_group"],
    "unit_price": ["商品原价", "单价", "unit_price"],
    "quantity": ["商品数量", "数量", "quantity"],
    "amount": ["实收金额", "金额", "订单金额", "amount"],
    "discount_amount": ["折让金额"],
    "remark": ["备注", "remark"],
    "logistics_no": ["物流单号", "物流号", "logistics_no"],
    "weight": ["重量", "weight"],
    "_pay_platform_no": ["支付平台单号"],
    "_takeout_platform_no": ["外卖平台单号"],
}

ORDER_LEVEL_FIELDS = {
    "order_no", "order_date", "store_name", "queue_no", "order_source",
    "delivery_method", "status", "payment_method", "member_no", "customer_name",
    "phone", "address", "remark",
    "_pay_platform_no", "_takeout_platform_no",
}
ROW_LEVEL_FIELDS = {
    "product_name", "flavor_group", "unit_price", "quantity", "amount",
    "discount_amount", "logistics_no", "weight",
}
NUMERIC_FIELDS = {"unit_price", "quantity", "amount", "discount_amount"}
DATE_FIELDS = {"order_date"}

INSERT_SQL = """
INSERT INTO coffee_orders (
  order_no, order_date, store_name, queue_no, order_source, delivery_method,
  status, payment_method, member_no, customer_name, phone, address,
  product_name, flavor_group, unit_price, quantity, amount,
  remark, logistics_no, weight, is_order_header,
  platform_order_no, discount_amount
) VALUES (
  %(order_no)s, %(order_date)s, %(store_name)s, %(queue_no)s, %(order_source)s, %(delivery_method)s,
  %(status)s, %(payment_method)s, %(member_no)s, %(customer_name)s, %(phone)s, %(address)s,
  %(product_name)s, %(flavor_group)s, %(unit_price)s, %(quantity)s, %(amount)s,
  %(remark)s, %(logistics_no)s, %(weight)s, %(is_order_header)s,
  %(platform_order_no)s, %(discount_amount)s
)
"""

_PLACEHOLDERS = {"", "-", "--", "none", "null", "n/a", "na"}


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        s = v.strip().lower()
        return s in _PLACEHOLDERS
    return False


def _normalize_text(v: Any) -> str | None:
    if _is_blank(v):
        return None
    return str(v).strip()


def _normalize_numeric(v: Any) -> Decimal | None:
    if _is_blank(v):
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    try:
        return Decimal(str(v).strip())
    except (ValueError, ArithmeticError):
        return None


def _normalize_date(v: Any) -> datetime | None:
    if _is_blank(v):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        try:
            return datetime.fromordinal(int(v) + 693594)
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _build_header_map(header_row: Any) -> dict[str, int]:
    """Scan header row, return {field: col_index} mapping by matching header
    names against FIELD_ALIASES. Columns whose headers don't match are ignored."""
    headers: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        raw = cell.value
        if raw is None:
            continue
        headers[idx] = str(raw).strip()

    field_to_col: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        for col, hname in headers.items():
            if hname in aliases:
                field_to_col[field] = col
                break
    return field_to_col


def _extract_values(row: Any, field_to_col: dict[str, int]) -> dict[str, Any]:
    """Read cell values from a row based on the field->col mapping."""
    cells = [cell.value for cell in row]
    values: dict[str, Any] = {}
    for field, col in field_to_col.items():
        values[field] = cells[col] if col < len(cells) else None
    return values


DEFAULT_STORE_NAME = "熊美式世博大道店"


def _build_record(values: dict[str, Any], context: dict[str, Any], is_header: bool) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for field in ORDER_LEVEL_FIELDS:
        if field.startswith("_"):
            continue
        record[field] = context.get(field)
    for field in ROW_LEVEL_FIELDS:
        raw = values.get(field)
        if field in NUMERIC_FIELDS:
            record[field] = _normalize_numeric(raw)
        elif field in DATE_FIELDS:
            record[field] = _normalize_date(raw)
        else:
            record[field] = _normalize_text(raw)
    # platform_order_no: order-level, prefer "支付平台单号", fallback to "外卖平台单号"
    pay_no = _normalize_text(context.get("_pay_platform_no"))
    takeout_no = _normalize_text(context.get("_takeout_platform_no"))
    record["platform_order_no"] = pay_no if pay_no else takeout_no
    # store_name default
    if not record.get("store_name"):
        record["store_name"] = DEFAULT_STORE_NAME
    record["is_order_header"] = 1 if is_header else 0
    return record


def _update_context(values: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    new_ctx = dict(context)
    for field in ORDER_LEVEL_FIELDS:
        raw = values.get(field)
        if field in DATE_FIELDS:
            new_ctx[field] = _normalize_date(raw)
        elif field in NUMERIC_FIELDS:
            new_ctx[field] = _normalize_numeric(raw)
        else:
            new_ctx[field] = _normalize_text(raw)
    return new_ctx


def _iter_records(file_bytes: bytes) -> Iterator[tuple[dict[str, Any], bool, bool]]:
    """Yield (record, is_header_row, is_empty_row) tuples."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        context: dict[str, Any] = {}
        row_iter = ws.iter_rows(values_only=False)

        try:
            header_row = next(row_iter)
        except StopIteration:
            return

        field_to_col = _build_header_map(header_row)
        if not field_to_col:
            raise ValueError("未能识别任何表头列，请检查 Excel 第一行是否为中文列名（订单号/日期/金额 等）")

        for row in row_iter:
            values = _extract_values(row, field_to_col)
            if all(_is_blank(v) for v in values.values()):
                yield {}, False, True
                continue
            order_no_raw = values.get("order_no")
            is_header = not _is_blank(order_no_raw)
            if not is_header:
                # Skip summary/total rows: order_no blank + no product detail
                prod = _normalize_text(values.get("product_name"))
                qty = _normalize_numeric(values.get("quantity"))
                if not prod and not qty:
                    continue
            if is_header:
                context = _update_context(values, context)
            record = _build_record(values, context, is_header)
            yield record, is_header, False
    finally:
        wb.close()


def import_excel(conn: Any, file_bytes: bytes) -> ImportResult:
    """Truncate coffee_orders and import all rows from the uploaded Excel bytes."""
    cursor = conn.cursor()
    errors: list[str] = []
    total_rows = 0
    skipped_empty = 0
    orders_seen: set[str] = set()
    stores_seen: set[str] = set()
    total_amount = Decimal("0")

    batch: list[dict[str, Any]] = []
    BATCH_SIZE = 500

    try:
        cursor.execute("TRUNCATE TABLE coffee_orders")

        for record, is_header, is_empty in _iter_records(file_bytes):
            if is_empty:
                skipped_empty += 1
                continue
            batch.append(record)
            total_rows += 1
            if is_header:
                on = record.get("order_no")
                if on:
                    orders_seen.add(on)
                st = record.get("store_name")
                if st:
                    stores_seen.add(st)
                amt = record.get("amount")
                if amt is not None:
                    total_amount += amt
            if len(batch) >= BATCH_SIZE:
                _flush(cursor, batch)
                batch.clear()

        if batch:
            _flush(cursor, batch)
            batch.clear()

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()

    return ImportResult(
        total_rows=total_rows,
        orders_count=len(orders_seen),
        total_amount=total_amount if total_amount else None,
        stores_count=len(stores_seen),
        skipped_empty_rows=skipped_empty,
        truncated=True,
        errors=errors,
    )


def _flush(cursor: Any, batch: list[dict[str, Any]]) -> None:
    if not batch:
        return
    cursor.executemany(INSERT_SQL, batch)
